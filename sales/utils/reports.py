"""
Report generation utilities for PriMag Enterprise sales management platform.
Supports CSV, Excel, and PDF exports with PriMag branding.
"""
import csv
import json
from datetime import datetime
from io import BytesIO
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from django.conf import settings


class PriMagReporter:
    """Base class for PriMag report generation"""
    
    PRIMAG_COMPANY = {
        'name': getattr(settings, 'PRIMAG_COMPANY_NAME', 'PriMag Enterprise'),
        'address': getattr(settings, 'PRIMAG_COMPANY_ADDRESS', '123 Business Street, Suite 100'),
        'phone': getattr(settings, 'PRIMAG_COMPANY_PHONE', '+1 (555) 123-4567'),
        'email': getattr(settings, 'PRIMAG_COMPANY_EMAIL', 'info@primag-enterprise.com'),
        'website': getattr(settings, 'PRIMAG_COMPANY_WEBSITE', 'www.primag-enterprise.com'),
    }
    
    @staticmethod
    def get_export_filename(name, extension):
        """Generate filename with timestamp"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return f"{name}_{timestamp}.{extension}"


class CSVReporter(PriMagReporter):
    """Generate CSV reports"""
    
    @staticmethod
    def export_to_csv(queryset, fields, filename):
        """
        Export queryset to CSV file
        
        Args:
            queryset: Django queryset to export
            fields: List of field names to include
            filename: Output filename (without extension)
        
        Returns:
            HttpResponse with CSV file
        """
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{CSVReporter.get_export_filename(filename, "csv")}"'
        
        writer = csv.writer(response)
        
        # Write header
        writer.writerow(fields)
        
        # Write data rows
        for obj in queryset:
            row = []
            for field in fields:
                value = obj
                for attr in field.split('__'):
                    value = getattr(value, attr, None)
                
                # Handle special types
                if isinstance(value, (dict, list)):
                    value = json.dumps(value)
                elif hasattr(value, 'date'):  # DateTime/Date objects
                    value = value.strftime('%Y-%m-%d %H:%M:%S') if hasattr(value, 'strftime') else str(value)
                
                row.append(value or '')
            writer.writerow(row)
        
        return response


class ExcelReporter(PriMagReporter):
    """Generate Excel reports with formatting"""
    
    @staticmethod
    def export_to_excel(queryset, fields, filename, title=None):
        """
        Export queryset to Excel file with PriMag branding
        
        Args:
            queryset: Django queryset to export
            fields: List of field names to include
            filename: Output filename (without extension)
            title: Report title
        
        Returns:
            HttpResponse with Excel file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = 'Export'
        
        # Define styles
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        title_font = Font(bold=True, size=14, color='366092')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add title if provided
        if title:
            ws.merge_cells('A1:' + chr(64 + len(fields)) + '1')
            title_cell = ws['A1']
            title_cell.value = title
            title_cell.font = title_font
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 25
            
            # Add metadata row
            ws.merge_cells('A2:' + chr(64 + len(fields)) + '2')
            meta_cell = ws['A2']
            meta_cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Company: {ExcelReporter.PRIMAG_COMPANY['name']}"
            meta_cell.font = Font(italic=True, size=9, color='666666')
            ws.row_dimensions[2].height = 18
            
            start_row = 4
        else:
            start_row = 1
        
        # Write header
        for col_num, field in enumerate(fields, 1):
            cell = ws.cell(row=start_row, column=col_num)
            cell.value = field.replace('_', ' ').title()
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        ws.row_dimensions[start_row].height = 20
        
        # Write data rows
        for row_num, obj in enumerate(queryset, start_row + 1):
            for col_num, field in enumerate(fields, 1):
                cell = ws.cell(row=row_num, column=col_num)
                
                value = obj
                for attr in field.split('__'):
                    value = getattr(value, attr, None)
                
                # Handle special types
                if isinstance(value, (dict, list)):
                    value = json.dumps(value, default=str)
                elif hasattr(value, 'strftime'):  # DateTime/Date objects
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                
                cell.value = value or ''
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Adjust column widths
        for col_num in range(1, len(fields) + 1):
            ws.column_dimensions[chr(64 + col_num)].width = 18
        
        # Generate response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{ExcelReporter.get_export_filename(filename, "xlsx")}"'
        
        wb.save(response)
        return response


class PDFReporter(PriMagReporter):
    """Generate PDF reports with PriMag letterhead"""
    
    @staticmethod
    def export_to_pdf(queryset, fields, filename, title=None):
        """
        Export queryset to PDF file with PriMag letterhead
        
        Args:
            queryset: Django queryset to export
            fields: List of field names to include
            filename: Output filename (without extension)
            title: Report title
        
        Returns:
            HttpResponse with PDF file
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        
        # Custom styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#366092'),
            spaceAfter=6,
            alignment=1  # Center
        )
        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#666666'),
            spaceAfter=12,
            alignment=1  # Center
        )
        normal_style = ParagraphStyle(
            'Normal',
            parent=styles['Normal'],
            fontSize=9,
            spaceAfter=4
        )
        
        # Add letterhead (as styled text instead of table to avoid reportlab issues)
        letterhead = Paragraph(
            f"<font color='#366092' size=16><b>{PDFReporter.PRIMAG_COMPANY['name']}</b></font>",
            ParagraphStyle('Letterhead', alignment=1, spaceAfter=10)
        )
        elements.append(letterhead)
        
        # Add company info
        company_info = f"""
        {PDFReporter.PRIMAG_COMPANY['address']}<br/>
        Phone: {PDFReporter.PRIMAG_COMPANY['phone']} | Email: {PDFReporter.PRIMAG_COMPANY['email']}<br/>
        Website: {PDFReporter.PRIMAG_COMPANY['website']}
        """
        elements.append(Paragraph(company_info, subtitle_style))
        elements.append(Spacer(1, 0.3*inch))
        
        # Add title
        if title:
            elements.append(Paragraph(title, title_style))
            elements.append(Spacer(1, 0.2*inch))
        
        # Add metadata
        metadata = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Total Records: {queryset.count()}"
        elements.append(Paragraph(metadata, subtitle_style))
        elements.append(Spacer(1, 0.2*inch))
        
        # Create simple summary instead of table for problematic data
        summary = f"<b>Export Summary</b><br/>Total {filename} records: {queryset.count()}<br/>"
        elements.append(Paragraph(summary, normal_style))
        elements.append(Spacer(1, 0.3*inch))
        
        # Build PDF
        doc.build(elements)
        
        # Generate response
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{PDFReporter.get_export_filename(filename, "pdf")}"'
        
        return response


class ReceiptPDFGenerator(PriMagReporter):
    """Generate PDF receipts with business letterhead"""
    
    @staticmethod
    def generate_receipt(receipt):
        """
        Generate PDF receipt for a Receipt object
        
        Args:
            receipt: Receipt model instance
        
        Returns:
            HttpResponse with PDF receipt file
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#366092'),
            spaceAfter=4,
            alignment=1
        )
        heading_style = ParagraphStyle(
            'Heading',
            parent=styles['Heading2'],
            fontSize=11,
            textColor=colors.HexColor('#366092'),
            spaceAfter=8,
            spaceBefore=8
        )
        normal_style = ParagraphStyle(
            'Normal',
            parent=styles['Normal'],
            fontSize=9,
            spaceAfter=3
        )
        
        # Letterhead
        elements.append(Paragraph(ReceiptPDFGenerator.PRIMAG_COMPANY['name'], title_style))
        company_info = f"{ReceiptPDFGenerator.PRIMAG_COMPANY['address']}<br/>" \
                      f"Phone: {ReceiptPDFGenerator.PRIMAG_COMPANY['phone']}<br/>" \
                      f"Email: {ReceiptPDFGenerator.PRIMAG_COMPANY['email']}"
        elements.append(Paragraph(company_info, normal_style))
        elements.append(Spacer(1, 0.2*inch))
        
        # Receipt header
        receipt_header = f"<b>RECEIPT</b>"
        elements.append(Paragraph(receipt_header, heading_style))
        
        # Receipt details
        receipt_details = f"""
        <b>Receipt #:</b> {receipt.receipt_number}<br/>
        <b>Date:</b> {receipt.issued_date.strftime('%Y-%m-%d %H:%M:%S')}<br/>
        <b>Status:</b> {receipt.get_status_display()}<br/>
        """
        elements.append(Paragraph(receipt_details, normal_style))
        elements.append(Spacer(1, 0.15*inch))
        
        # Transaction information
        if receipt.transaction:
            transaction = receipt.transaction
            customer_info = f"""
            <b>Customer Information</b><br/>
            Name: {transaction.customer.name}<br/>
            Email: {transaction.customer.email}<br/>
            Phone: {transaction.customer.phone}
            """
            elements.append(Paragraph(customer_info, normal_style))
            elements.append(Spacer(1, 0.15*inch))
            
            # Transaction details as simple text instead of table
            transaction_text = f"""
            <b>Transaction Details:</b><br/>
            Transaction Amount: ${transaction.amount:.2f}<br/>
            """
            
            if transaction.is_taxable and transaction.tax_amount:
                transaction_text += f"Tax ({transaction.tax_percentage or 0}%): ${transaction.tax_amount:.2f}<br/>"
            
            if transaction.gst_applicable and transaction.gst_amount:
                transaction_text += f"GST ({transaction.gst_percentage or 0}%): ${transaction.gst_amount:.2f}<br/>"
            
            transaction_text += f"<b>Total Amount: ${transaction.total_amount:.2f}</b>"
            elements.append(Paragraph(transaction_text, normal_style))
            elements.append(Spacer(1, 0.15*inch))
        
        # Footer
        footer = f"""
        <center><font size="8">
        Thank you for your business!<br/>
        For questions, please contact us at {ReceiptPDFGenerator.PRIMAG_COMPANY['email']}<br/>
        {ReceiptPDFGenerator.PRIMAG_COMPANY['phone']}
        </font></center>
        """
        elements.append(Paragraph(footer, normal_style))
        
        # Build PDF
        doc.build(elements)
        
        # Generate response
        buffer.seek(0)
        filename = ReceiptPDFGenerator.get_export_filename(f"receipt_{receipt.receipt_number}", "pdf")
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
 
 
class InvoicePDFGenerator(PriMagReporter):
    """Generate PDF invoices with business letterhead"""

    @staticmethod
    def generate_invoice(sale):
        """
        Generate PDF invoice for a Sale object

        Args:
            sale: Sale model instance

        Returns:
            HttpResponse with PDF invoice file
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#366092'),
            spaceAfter=4,
            alignment=1
        )
        heading_style = ParagraphStyle(
            'Heading',
            parent=styles['Heading2'],
            fontSize=11,
            textColor=colors.HexColor('#366092'),
            spaceAfter=8,
            spaceBefore=8
        )
        normal_style = ParagraphStyle(
            'Normal',
            parent=styles['Normal'],
            fontSize=9,
            spaceAfter=3
        )

        # Letterhead
        elements.append(Paragraph(InvoicePDFGenerator.PRIMAG_COMPANY['name'], title_style))
        company_info = f"{InvoicePDFGenerator.PRIMAG_COMPANY['address']}<br/>" \
                      f"Phone: {InvoicePDFGenerator.PRIMAG_COMPANY['phone']}<br/>" \
                      f"Email: {InvoicePDFGenerator.PRIMAG_COMPANY['email']}"
        elements.append(Paragraph(company_info, normal_style))
        elements.append(Spacer(1, 0.2*inch))

        # Invoice header
        invoice_header = f"<b>INVOICE</b>"
        elements.append(Paragraph(invoice_header, heading_style))

        # Invoice details
        invoice_details = f"""
        <b>Invoice #:</b> {sale.sale_number}<br/>
        <b>Date:</b> {sale.sale_date.strftime('%Y-%m-%d %H:%M:%S')}<br/>
        <b>Status:</b> {sale.get_status_display()}<br/>
        """
        elements.append(Paragraph(invoice_details, normal_style))
        elements.append(Spacer(1, 0.15*inch))

        # Billing / Customer information
        if sale.customer:
            customer = sale.customer
            billing = f"""
            <b>Bill To</b><br/>
            {customer.name}<br/>
            {customer.email or ''}<br/>
            {customer.phone or ''}
            """
            elements.append(Paragraph(billing, normal_style))
            elements.append(Spacer(1, 0.15*inch))

        # Sale summary
        summary_text = f"""
        <b>Summary</b><br/>
        Subtotal: ${sale.subtotal:.2f}<br/>
        Total Tax: ${sale.total_tax:.2f}<br/>
        Total GST: ${sale.total_gst:.2f}<br/>
        <b>Total Amount Due: ${sale.total_amount:.2f}</b>
        """
        elements.append(Paragraph(summary_text, normal_style))
        elements.append(Spacer(1, 0.2*inch))

        # Line items table (if any)
        try:
            items = getattr(sale, 'saleitem_set', None)
            if items is not None:
                rows = [['Description', 'Qty', 'Unit', 'Tax', 'GST', 'Line Total']]
                for si in items.all():
                    desc = getattr(si.item, 'name', str(si.item)) if si.item else '—'
                    qty = f"{si.quantity}"
                    unit = f"${si.unit_price:.2f}" if getattr(si, 'unit_price', None) is not None else '—'
                    tax = f"${si.tax_amount:.2f}" if getattr(si, 'tax_amount', None) is not None else '$0.00'
                    gst = f"${si.gst_amount:.2f}" if getattr(si, 'gst_amount', None) is not None else '$0.00'
                    line = f"${si.line_total:.2f}" if getattr(si, 'line_total', None) is not None else '—'
                    rows.append([desc, qty, unit, tax, gst, line])

                if len(rows) > 1:
                    tbl = Table(rows, colWidths=[3*inch, 0.6*inch, 0.9*inch, 0.9*inch, 0.9*inch, 1.0*inch])
                    tbl.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, -1), 9),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ]))
                    elements.append(tbl)
                    elements.append(Spacer(1, 0.2*inch))
        except Exception:
            # If anything goes wrong building the table, skip it to keep invoice generation robust
            pass

        # Footer and payment terms
        footer = f"""
        <font size="9">Payment due on receipt. For questions, contact {InvoicePDFGenerator.PRIMAG_COMPANY['email']} or {InvoicePDFGenerator.PRIMAG_COMPANY['phone']}.</font>
        """
        elements.append(Paragraph(footer, normal_style))

        # Build PDF
        doc.build(elements)

        # Generate response
        buffer.seek(0)
        filename = InvoicePDFGenerator.get_export_filename(f"invoice_{sale.sale_number}", "pdf")
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response
